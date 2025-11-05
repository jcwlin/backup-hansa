import os
import tempfile
import pickle
from datetime import datetime
import pandas as pd
import yaml
from tabulate import tabulate
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify, session, send_from_directory
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash

# 移除多語系支持
# from translations import get_text, get_language_options
from analyzers.clients.gemini_client import call_gemini as call_gemini_api
from langdetect import detect, DetectorFactory
import pdfplumber
import docx
from PIL import Image
import os
import pytesseract
from bs4 import BeautifulSoup
import email
import email.policy
import re, json
import concurrent.futures
import threading
import uuid
from pdf2image import convert_from_path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor
import logging
import sqlite3
from typing import List, Dict, Any, Optional

DATA_DIR = '/data'
os.makedirs(DATA_DIR, exist_ok=True)

# Use /data for all files that must persist
HISTORY_FILE = os.path.join(DATA_DIR, 'history.pkl')
DB_PATH = os.path.join(DATA_DIR, 'app.db')
UPLOAD_FOLDER = os.path.join(DATA_DIR, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

import subprocess
try:
    result = subprocess.run(["pdftotext", "-v"], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    print("Poppler check:", result.stderr.decode() or result.stdout.decode())
except Exception as e:
    print("Poppler not found in PATH:", e)

logging.basicConfig(level=logging.INFO)

# Check current directory and files
logging.info(f"Current working directory: {os.getcwd()}")
logging.info(f"Files in project root: {os.listdir('.')}")
if os.path.exists('data'):
    logging.info(f"Files in 'data' folder: {os.listdir('data')}")
else:
    logging.warning("'data' folder does not exist!")



# import platform
#
# if platform.system() == 'Windows':
#     poppler_path = r'C:\Program Files\poppler-24.08.0\Library\bin'
#     if os.path.exists(poppler_path):
#         os.environ['PATH'] += os.pathsep + poppler_path
#
#     # Configure Tesseract path for OCR
#     tesseract_path = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
#     if os.path.exists(tesseract_path):
#         pytesseract.pytesseract.tesseract_cmd = tesseract_path
# 設置日誌配置
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

DetectorFactory.seed = 0


app = Flask(__name__)
app.secret_key = 'fileanalyzer_secret_key'

# 移除多語系模板過濾器
# @app.template_filter('t')
# def translate_filter(key, **kwargs):
#     """模板中的翻譯過濾器"""
#     lang = session.get('language', 'zh')
#     return get_text(key, lang, **kwargs)

# @app.context_processor
# def inject_language():
#     """注入語言相關變數到所有模板"""
#     return {
#         'current_language': session.get('language', 'zh'),
#         'language_options': get_language_options()
#     }

# Set upload folder
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER, mode=0o755)
    app.logger.info(f"Created upload folder: {UPLOAD_FOLDER}")
# 確保上傳目錄有正確權限
try:
    os.chmod(UPLOAD_FOLDER, 0o755)
    app.logger.info(f"Set permissions for upload folder: {UPLOAD_FOLDER}")
except Exception as e:
    app.logger.warning(f"Failed to set permissions for upload folder: {e}")

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB 限制

HISTORY_FILE = 'history.pkl'

# 讀取 config.yaml
with open('config.yaml', 'r', encoding='utf-8') as f:
    config = yaml.safe_load(f)

# 導入 prompts 從 analyzers/services/prompts
try:
    from analyzers.services.prompts import SERVICE_PROMPTS
    # 將 SERVICE_PROMPTS 合併到 config 中
    config['prompts'] = SERVICE_PROMPTS
except ImportError as e:
    print(f"Warning: Could not import SERVICE_PROMPTS: {e}")
    # 如果導入失敗，保持原有的 config['prompts']

# 移除多語系支持，只保留英文
# LANGS = list(config['langs'].keys())

# History records
if os.path.exists(HISTORY_FILE):
    with open(HISTORY_FILE, 'rb') as f:
        history = pickle.load(f)
else:
    history = []

# 若有安裝 docling，可用於更強大的文件解析
try:
    from docling.document_converter import DocumentConverter
    DOC_CONVERTER = DocumentConverter()
except ImportError:
    DOC_CONVERTER = None

# Note: Moved to after init_db definition to avoid static analyzer false positive order issues

# Progress tracking
progress_store = {}
progress_lock = threading.Lock()

# Global variable for storing VESSEL and VOY.NO. information
vessel_voy_data = None

def update_progress(task_id, current, total, message=None):
    """更新任務進度"""
    try:
        percentage = (current / total * 100) if total > 0 else 0
        with progress_lock:
            progress_store[task_id] = {
                'current': current,
                'total': total,
                'percentage': percentage,
                'message': message or f'Completed {current}/{total} files'
            }
    except Exception as e:
        app.logger.error(f"Error updating progress for task {task_id}: {str(e)}")

def get_progress(task_id):
    """取得任務進度"""
    try:
        with progress_lock:
            progress = progress_store.get(task_id, {
                'current': 0,
                'total': 1,
                'percentage': 0,
            })
            
            # 記錄進度請求
            app.logger.info(f"Progress request for task {task_id}: has_final_result={'final_result' in progress}")
            
            # 如果任務完成且有最終結果，返回最終結果
            if 'final_result' in progress:
                app.logger.info(f"Returning final result for task {task_id}: {progress['final_result'].get('success', 'unknown')}")
                return jsonify(progress['final_result'])
            
            return jsonify(progress)
    except Exception as e:
        app.logger.error(f"Error getting progress for task {task_id}: {str(e)}")
        return jsonify({
            'current': 0,
            'total': 1,
            'percentage': 0,
            'message': f'Error retrieving progress: {str(e)}'
        }), 500

# =========================
# User and Permission Management: SQLite Simple Management
# =========================
DB_PATH = os.path.join(os.path.dirname(__file__), 'app.db')

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def get_user_by_username(username):
    """Get user information by username"""
    conn = get_db_connection()
    try:
        user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        if user:
            return dict(user)
        return None
    except Exception as e:
        logging.error(f"Failed to get user information: {e}")
        return None
    finally:
        conn.close()

def init_db():
    """Initialize database tables and ensure admin user exists."""
    import logging
    from werkzeug.security import generate_password_hash

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Create users table if it doesn't exist
        cur.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                is_admin INTEGER NOT NULL DEFAULT 0,
                is_active INTEGER NOT NULL DEFAULT 1
            )
        ''')
        # NEW: History table
        cur.execute('''
                    CREATE TABLE IF NOT EXISTS history (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        username TEXT NOT NULL,
                        timestamp TEXT NOT NULL,
                        files TEXT NOT NULL,
                        analysis_type TEXT,
                        lang TEXT,
                        rows INTEGER,
                        cols INTEGER,
                        headers TEXT,
                        tokens INTEGER,
                        excel TEXT,
                        log_file TEXT,
                        time_cost REAL,
                        duration_str TEXT,
                        total_pages INTEGER,
                        saved_files TEXT,
                        FOREIGN KEY(username) REFERENCES users(username)
                    )
                ''')

        # Create user_analyzers table if it doesn't exist
        cur.execute('''
            CREATE TABLE IF NOT EXISTS user_analyzers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                analyzer TEXT NOT NULL,
                enabled INTEGER NOT NULL DEFAULT 1,
                vlm_provider TEXT NOT NULL DEFAULT 'cloud',
                ollama_model TEXT DEFAULT NULL,
                ocr_lang TEXT DEFAULT 'auto',
                save_files INTEGER NOT NULL DEFAULT 0,
                UNIQUE(user_id, analyzer),
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        ''')

        # Add "save_files" column if it doesn't exist
        try:
            cur.execute("ALTER TABLE user_analyzers ADD COLUMN save_files INTEGER NOT NULL DEFAULT 0")
        except Exception:
            pass  # Ignore if it already exists

        # Ensure default admin user exists (avoid duplicate insertion)
        cur.execute("SELECT id FROM users WHERE username = ?", ('admin',))
        if cur.fetchone() is None:
            hashed_pw = generate_password_hash('admin123', method='pbkdf2:sha256')
            cur.execute(
                "INSERT INTO users (username, password_hash, is_admin, is_active) VALUES (?, ?, ?, ?)",
                ('admin', hashed_pw, 1, 1)
            )
            print("[DB] Default admin created: username='admin', password='admin123'")
        else:
            print("[DB] Admin user already exists — skipping creation.")

        conn.commit()
        conn.close()

    except Exception as e:
        logging.error(f"Failed to initialize database: {e}")
def migrate_db_schema():
    """Add missing columns to existing tables (safe migration)."""
    conn = get_db_connection()
    cur = conn.cursor()

    new_columns = {
        "display_name": "TEXT",
        "email": "TEXT",
        "phone": "TEXT",
        "address": "TEXT",
        "logo_file": "TEXT",
        "notes": "TEXT"
    }

    for col, col_type in new_columns.items():
        try:
            cur.execute(f"ALTER TABLE users ADD COLUMN {col} {col_type}")
            print(f"✅ Added column: {col}")
        except sqlite3.OperationalError:
            pass

    conn.commit()
    conn.close()
    print("✅ Database migration complete.")


# Initialize database at startup (init_db already defined here)
try:
    init_db()
    migrate_db_schema()
except Exception as e:
    logging.error(f"Failed to initialize database: {e}")

def get_current_user():
    uid = session.get('user_id')
    if not uid:
        return None
    conn = get_db_connection()
    row = conn.execute('SELECT * FROM users WHERE id=?', (uid,)).fetchone()
    conn.close()
    return row

def calculate_user_statistics(history_records):
    """Calculate user statistics including total tokens and daily usage"""
    from collections import defaultdict
    from datetime import datetime
    import re
    
    user_stats = defaultdict(lambda: {
        'total_tokens': 0,
        'daily_usage': defaultdict(int),
        'total_records': 0,
        'total_files': 0,
        'total_processed': 0
    })
    
    for record in history_records:
        username = record.get('username', 'Unknown User')
        tokens = record.get('tokens', 0)
        
        if isinstance(tokens, (int, float)):
            tokens = int(tokens)
        else:
            tokens = 0
            
        user_stats[username]['total_tokens'] += tokens
        user_stats[username]['total_records'] += 1
        
        # Count files
        files = record.get('files', [])
        if files:
            user_stats[username]['total_files'] += len(files)
        
        # Count processing results (extract row count from result field)
        result_str = record.get('result', '')
        if result_str:
            # Try to extract row count from result, e.g. "24 rows 13 columns" -> 24
            match = re.search(r'(\d+)行', result_str)
            if match:
                rows = int(match.group(1))
                user_stats[username]['total_processed'] += rows
            else:
                # If no row count info, calculate by file count
                user_stats[username]['total_processed'] += len(files) if files else 1
        
        # Parse date
        time_str = record.get('time') or record.get('timestamp', '')
        if time_str:
            try:
                # Try to parse date part (YYYY-MM-DD)
                if ' ' in time_str:
                    date_part = time_str.split(' ')[0]
                else:
                    date_part = time_str
                
                # Validate date format
                datetime.strptime(date_part, '%Y-%m-%d')
                user_stats[username]['daily_usage'][date_part] += tokens
            except ValueError:
                # Skip if date format is invalid
                pass
    
    return dict(user_stats)

def create_bilingual_log(log_entries_zh, log_filename):
    """Create English-Chinese bilingual log file (English version first, Chinese version second)"""
    import re
    
    # English-Chinese translation mapping
    translations = {
        '=== 檔案分析處理日誌 ===': '=== File Analysis Processing Log ===',
        '開始時間': 'Start Time',
        '分析類型': 'Analysis Type',
        '語言': 'Language',
        '檔案數量': 'File Count',
        '自定義提示詞': 'Custom Prompt',
        '是': 'Yes',
        '否': 'No',
        '=== OCR 文字提取階段 ===': '=== OCR Text Extraction Phase ===',
        '檔案準備': 'File Preparation',
        'OCR 提取': 'OCR Extraction',
        '頁數': 'Pages',
        '提取文字長度': 'Extracted Text Length',
        '字符': 'characters',
        '狀態': 'Status',
        '成功': 'Success',
        '失敗': 'Failed',
        '=== LLM 分析階段 ===': '=== LLM Analysis Phase ===',
        'LLM 分析': 'LLM Analysis',
        'Tokens 使用': 'Tokens Used',
        '分析結果': 'Analysis Result',
        'VLM 補救欄位': 'VLM Remediated Fields',
        'VLM Tokens 使用': 'VLM Tokens Used',
        'VLM 補救': 'VLM Remediation',
        '無需要': 'Not Required',
        '=== 處理摘要 ===': '=== Processing Summary ===',
        '總處理時間': 'Total Processing Time',
        '秒': 'seconds',
        '總 Tokens 使用': 'Total Tokens Used',
        '總頁數': 'Total Pages',
        '成功處理檔案': 'Successfully Processed Files',
        '=== 詳細檔案處理統計 ===': '=== Detailed File Processing Statistics ===',
        '檔案': 'File',
        'LLM Tokens': 'LLM Tokens',
        'VLM Tokens': 'VLM Tokens',
        '總 Tokens': 'Total Tokens',
        '成功提取欄位': 'Successfully Extracted Fields',
        '缺失欄位': 'Missing Fields',
        'VLM 補救欄位': 'VLM Remediated Fields',
        '結束時間': 'End Time',
        '輸出 Excel': 'Output Excel',
        '輸出 Log': 'Output Log',
        '個': '',
        ' bytes': ' bytes'
    }
    
    # Special pattern translations (for handling compound words)
    pattern_translations = [
        (r'成功提取欄位', 'Successfully Extracted Fields'),
        (r'成功處理檔案', 'Successfully Processed Files'),
        (r'總([A-Za-z\s]+):', r'Total \1:'),
        (r'成功([A-Za-z\s]+)', r'Successfully \1'),
        (r'檔案([A-Za-z\s]*)', r'File\1'),
        (r'([A-Za-z\s]+)欄位', r'\1 Fields'),
        (r'([A-Za-z\s]+)頁數', r'\1 Pages'),
        (r'([A-Za-z\s]+)分析', r'\1 Analysis')
    ]
    
    # Create English version of log entries
    log_entries_en = []
    for entry in log_entries_zh:
        en_entry = entry
        
        # First perform basic translation
        for zh_text, en_text in translations.items():
            if zh_text in en_entry:
                en_entry = en_entry.replace(zh_text, en_text)
        
        # Then perform pattern translation
        for pattern, replacement in pattern_translations:
            en_entry = re.sub(pattern, replacement, en_entry)
        
        # Handle special Chinese-English mixed cases
        # Remove remaining Chinese characters but keep numbers and symbols
        def clean_mixed_text(text):
            # Protect Chinese characters in filenames
            if '.pdf' in text or '.xlsx' in text:
                return text
            
            # Special handling for statistics rows
            if '  - ' in text or '    * ' in text:
                # Separate indentation, Chinese part and English/number part
                indent_match = re.match(r'^(\s*[-*]\s*)', text)
                if indent_match:
                    indent = indent_match.group(1)
                    content = text[len(indent):]
                    
                    # If content contains Chinese, try to clean further
                    if re.search(r'[\u4e00-\u9fff]', content):
                        # Keep numbers, English, colons, spaces, etc.
                        cleaned_content = re.sub(r'[\u4e00-\u9fff]+', '', content)
                        # Clean up extra spaces
                        cleaned_content = re.sub(r'\s+', ' ', cleaned_content).strip()
                        if cleaned_content:
                            return indent + cleaned_content
                    return text
            
            return text
        
        en_entry = clean_mixed_text(en_entry)
        log_entries_en.append(en_entry)
    
    # Build bilingual log: English version first, Chinese version second
    bilingual_log = []
    
    # English version
    bilingual_log.append("=" * 80)
    bilingual_log.append("ENGLISH VERSION")
    bilingual_log.append("=" * 80)
    bilingual_log.extend(log_entries_en)
    
    bilingual_log.append("")
    bilingual_log.append("")
    
    # Chinese version
    bilingual_log.append("=" * 80)
    bilingual_log.append("中文版本")
    bilingual_log.append("=" * 80)
    bilingual_log.extend(log_entries_zh)
    
    return bilingual_log

def login_required(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user:
            return redirect(url_for('login', next=request.path))
        if user['is_active'] != 1:
            session.clear()
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapper

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user or user['is_admin'] != 1:
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return wrapper

def get_user_analyzer_setting(user_id: int, analysis_type: str):
    conn = get_db_connection()
    row = conn.execute(
        'SELECT enabled, vlm_provider, ollama_model, ocr_lang, save_files FROM user_analyzers WHERE user_id=? AND analyzer=?',
        (user_id, analysis_type)
    ).fetchone()
    conn.close()
    if not row:
        # If not set, default to enabled and use cloud
        return {'enabled': True, 'vlm_provider': 'cloud', 'ollama_model': None, 'ocr_lang': 'auto', 'save_files': False}
    return {'enabled': bool(row['enabled']), 'vlm_provider': row['vlm_provider'], 'ollama_model': row['ollama_model'], 'ocr_lang': row['ocr_lang'] or 'auto', 'save_files': bool(row['save_files'])}

def get_user_allowed_analyzers(user_id: int):
    # Return list of analyzers enabled by user. If no settings, default to allow all prompts keys in config
    conn = get_db_connection()
    rows = conn.execute('SELECT analyzer FROM user_analyzers WHERE user_id=? AND enabled=1', (user_id,)).fetchall()
    conn.close()
    if not rows:
        return list(config.get('prompts', {}).keys())
    return [r['analyzer'] for r in rows]

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        conn = get_db_connection()
        user = conn.execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
        conn.close()
        if user and check_password_hash(user['password_hash'], password) and user['is_active'] == 1:
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['is_admin'] = user['is_admin']
            next_url = request.args.get('next') or url_for('index')
            return redirect(next_url)
        return render_template('login.html', error='Invalid username or password, or account is disabled')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

def save_history():
    with open(HISTORY_FILE, 'wb') as f:
        pickle.dump(history, f)

def extract_json_from_response(resp):
    resp = resp.strip()
    # Remove markdown markers
    if resp.startswith('```json'):
        resp = resp[7:]
    if resp.startswith('```'):
        resp = resp[3:]
    if resp.endswith('```'):
        resp = resp[:-3]
    resp = resp.strip()
    # Try direct parsing
    try:
        return json.loads(resp)
    except Exception:
        pass
    # Try to extract array
    match = re.search(r'\[.*\]', resp, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(0))
        except Exception:
            pass
    # Try to extract first object
    match = re.search(r'\{[\s\S]*\}', resp)
    if match:
        try:
            return json.loads(match.group(0))
        except Exception:
            pass
    # If all fail, return original but provide clearer error message
    return {
        'error': 'Gemini API response is not in JSON format',
        'response': resp[:500] + '...' if len(resp) > 500 else resp
    }

def extract_text_from_pdf(file_path):
    """Extract text from PDF using pdfplumber, fallback to OCR if needed"""
    import pdfplumber
    import pytesseract
    from pdf2image import convert_from_path
    from PIL import Image
    import io, os

    text = ""

    try:
        # Try pdfplumber first (works for text-based PDFs)
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except Exception as e:
        print(f"[extract_text_from_pdf][ERROR] pdfplumber failed: {e}")

    # If no text, try OCR (for scanned PDFs)
    if not text.strip():
        print(f"[extract_text_from_pdf][INFO] No text found in {os.path.basename(file_path)}, running OCR...")
        try:
            images = convert_from_path(file_path)
            for img in images:
                ocr_text = pytesseract.image_to_string(img)
                text += ocr_text + "\n"
        except Exception as e:
            print(f"[extract_text_from_pdf][ERROR] OCR failed: {e}")

    return text.strip() if text else None


def extract_text_from_word(file_path):
    """Extract text from Word document"""
    try:
        doc = docx.Document(file_path)
        texts = []
        for paragraph in doc.paragraphs:
            texts.append(paragraph.text)
        return '\n'.join(texts)
    except Exception as e:
        app.logger.error(f"Word extraction error {file_path}: {str(e)}")
        return None

def extract_text_from_excel(file_path):
    """Extract text from Excel"""
    try:
        df = pd.read_excel(file_path)
        return df.to_string()
    except Exception as e:
        app.logger.error(f"Excel extraction error {file_path}: {str(e)}")
        return None

def extract_text_from_csv(file_path):
    """Extract text from CSV"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        app.logger.error(f"CSV extraction error {file_path}: {str(e)}")
        return None

def extract_text_from_image(file_path):
    """Extract text from image (OCR)"""
    try:
        image = Image.open(file_path)
        ocr_lang = config.get('ocr_lang', 'eng') or 'eng'
        text = pytesseract.image_to_string(image, lang=ocr_lang)
        return text
    except Exception as e:
        app.logger.error(f"Image OCR error {file_path}: {str(e)}")
        return None

def extract_file_content(file_path):
    """Extract file content"""
    try:
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == '.pdf':
            return extract_text_from_pdf(file_path)
        elif file_ext in ['.doc', '.docx']:
            return extract_text_from_word(file_path)
        elif file_ext in ['.xls', '.xlsx']:
            return extract_text_from_excel(file_path)
        elif file_ext == '.csv':
            return extract_text_from_csv(file_path)
        elif file_ext == '.txt':
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        elif file_ext in ['.jpg', '.jpeg', '.png', '.bmp']:
            return extract_text_from_image(file_path)
        else:
            app.logger.warning(f"Unsupported file type: {file_ext}")
            return None
            
    except Exception as e:
        app.logger.error(f"Error occurred while extracting file content {file_path}: {str(e)}")
        return None

from analyzers.services import analyze_text_with_prompt_with_gemini

def export_to_excel(data, filename, preview_title=None, manual_data=None, include_logo=False, user_info=None, vessel_voy_info=None, keep_filename=False):
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as OpenpyxlImage
        
        # 數據已經在 cargo_bl_postprocess 中處理過了，直接使用
        processed_data = data
        app.logger.info(f"export_to_excel: keep_filename={keep_filename}")
        app.logger.info(f"export_to_excel: data keys sample={[list(item.keys()) if isinstance(item, dict) else 'not_dict' for item in data[:2]]}")
        
        df = pd.DataFrame(processed_data)
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        df.to_excel(excel_path, index=False)
        wb = load_workbook(excel_path)
        ws = wb.active
        current_row = 1
        
        # 1. 標題和 Logo 同行
        if preview_title:
            today_str = datetime.now().strftime('%Y-%m-%d')
            
            # 插入標題行
            ws.insert_rows(current_row, 1)
            
            # 設置標題行的淡藍色底色
            light_blue_fill = PatternFill('solid', fgColor='E6F3FF')
            
            # 標題文字（左邊和中間）
            if ws.max_column > 1:
                # 合併前幾欄給標題
                merge_end_col = max(1, ws.max_column - 2) if include_logo else ws.max_column
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=merge_end_col)
            
            title_cell = ws.cell(row=current_row, column=1)
            # 使用用戶名稱，如果沒有則使用默認標題
            if user_info and user_info.get('display_name'):
                title_cell.value = user_info['display_name']
            else:
                title_cell.value = preview_title
            title_cell.font = Font(size=20, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.fill = light_blue_fill
            
            # Logo（右邊）
            if include_logo:
                # 優先使用用戶的logo，如果沒有則使用默認logo
                logo_path = None
                if user_info and user_info.get('logo_file'):
                    user_logo_path = os.path.join(app.config['UPLOAD_FOLDER'], user_info['logo_file'])
                    if os.path.exists(user_logo_path):
                        logo_path = user_logo_path
                
                # 如果沒有用戶logo，使用默認logo
                if not logo_path:
                    default_logo_path = os.path.join('static', 'hansa.png')
                    if os.path.exists(default_logo_path):
                        logo_path = default_logo_path
                
                if logo_path:
                    img = OpenpyxlImage(logo_path)
                    img.width = 60
                    img.height = 60
                    # 設置行高以容納 logo
                    ws.row_dimensions[current_row].height = 50
                    # 將 logo 放在最後一欄
                    last_col_letter = get_column_letter(ws.max_column)
                    ws.add_image(img, f'{last_col_letter}{current_row}')
                    
                    # 為 logo 所在的欄位也設置底色
                    logo_cell = ws.cell(row=current_row, column=ws.max_column)
                    logo_cell.fill = light_blue_fill
            
            # 為整行設置底色
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=current_row, column=col)
                if cell.fill.fgColor.rgb != light_blue_fill.fgColor.rgb:
                    cell.fill = light_blue_fill
            
            current_row += 1
            
            # 日期行
            ws.insert_rows(current_row, 1)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=ws.max_column)
            date_cell = ws.cell(row=current_row, column=1)
            date_cell.value = today_str
            date_cell.font = Font(size=12, italic=True)
            date_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # VESSEL 和 VOY.NO. 行（針對 Cargo_BL）
            if vessel_voy_info:
                # VESSEL 行
                ws.insert_rows(current_row, 1)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=ws.max_column)
                vessel_cell = ws.cell(row=current_row, column=1)
                vessel_cell.value = f"VESSEL: {vessel_voy_info.get('VESSEL', '')}"
                vessel_cell.font = Font(size=12, bold=True)
                vessel_cell.alignment = Alignment(horizontal='left', vertical='center')
                current_row += 1
                
                # VOY.NO. 行
                ws.insert_rows(current_row, 1)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=ws.max_column)
                voy_cell = ws.cell(row=current_row, column=1)
                voy_cell.value = f"VOY.NO.: {vessel_voy_info.get('VOY.NO.', '')}"
                voy_cell.font = Font(size=12, bold=True)
                voy_cell.alignment = Alignment(horizontal='left', vertical='center')
                current_row += 1
        
        # 2. 手動插入的資料（處理結構化數據）
        if manual_data and manual_data.get('rows'):
            manual_rows = manual_data['rows']
            if manual_rows:
                rows_to_insert = len(manual_rows)
                ws.insert_rows(current_row, rows_to_insert)
                
                for row_idx, row_data in enumerate(manual_rows):
                    for col_idx, cell_data in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row + row_idx, column=col_idx)
                        
                        # 處理結構化的 cell_data
                        if isinstance(cell_data, dict):
                            cell.value = cell_data.get('value', '')
                            color = cell_data.get('color', '#000000')
                            font_size = cell_data.get('fontSize', 12)
                        else:
                            cell.value = cell_data
                            color = '#000000'
                            font_size = 12
                        
                        # 設置字體樣式
                        cell.font = Font(bold=True, size=font_size, color=color.replace('#', ''))
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                current_row += rows_to_insert
        
        # 3. 主要資料的標題行樣式
        header_row = current_row
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='000000')
        for cell in ws[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 4. 欄寬自動調整（根據內容和字體大小）
        for col, cell in enumerate(ws[header_row], 1):
            # 計算該欄的最大內容長度
            max_len = max(len(str(cell.value)),
                          max((len(str(ws.cell(row=row, column=col).value)) for row in range(header_row+1, ws.max_row+1)), default=0))
            
            # 計算該欄的最大字體大小，處理 None 值
            font_sizes = []
            for row in range(1, ws.max_row+1):
                font_size = getattr(ws.cell(row=row, column=col).font, 'size', 12)
                if font_size is not None:
                    font_sizes.append(font_size)
                else:
                    font_sizes.append(12)
            max_font = max(font_sizes, default=12)
            
            # 根據字體大小調整欄寬
            adjusted_width = max(10, min(max_len * max_font / 12 + 2, 60))
            ws.column_dimensions[get_column_letter(col)].width = adjusted_width
        
        # 5. Total row 和 Cgos row 樣式
        for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
            first_cell = row[0].value
            if str(first_cell).strip().lower() == 'total':
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill('solid', fgColor='E5C29F')
            if str(first_cell).strip().startswith('# Cgos'):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill('solid', fgColor='C6EFCE')
        
        wb.save(excel_path)
        return excel_path
    except Exception as e:
        app.logger.error(f"匯出 Excel 錯誤: {str(e)}")
        return None

def save_to_history(entry):
    try:
        import json
        conn = get_db_connection()
        conn.execute('''
            INSERT INTO history (username, timestamp, files, analysis_type, lang, rows, cols, 
                                headers, tokens, excel, log_file, time_cost, duration_str, 
                                total_pages, saved_files)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            entry.get('username'),
            entry.get('time'),
            json.dumps(entry.get('files', [])),
            entry.get('analysis_type'),
            entry.get('lang'),
            entry.get('rows'),
            entry.get('cols'),
            json.dumps(entry.get('headers', [])),
            entry.get('tokens'),
            entry.get('excel'),
            entry.get('log_file'),
            entry.get('time_cost'),
            entry.get('duration_str'),
            entry.get('total_pages'),
            json.dumps(entry.get('saved_files', []))
        ))
        conn.commit()
        conn.close()
    except Exception as e:
        app.logger.error(f"Failed to save history: {str(e)}")

# 修正檔案內容提取函數名稱
def extract_text(file_path, file_extension):
    """提取檔案文字內容"""
    return extract_file_content(file_path)

@app.route('/', methods=['GET'])
@login_required
def index():
    lang = 'en'  # Default to English
    user = get_current_user()
    allowed = get_user_allowed_analyzers(user['id']) if user else list(config.get('prompts', {}).keys())
    # Filter config['prompts'] to show only user-available analyzers
    filtered_config = dict(config)
    filtered_config['prompts'] = {k: v for k, v in config.get('prompts', {}).items() if k in allowed}
    first_prompt_key = next(iter(filtered_config['prompts'].keys())) if filtered_config['prompts'] else ''
    return render_template('index.html', config=filtered_config, lang=lang, 
                         first_prompt_key=first_prompt_key, allowed_analyzers=allowed)

@app.route('/analyze', methods=['POST'])
@login_required
def analyze():
    try:
        # 生成唯一任務ID
        task_id = str(uuid.uuid4())
        
        files = request.files.getlist('files')
        analysis_type = request.form.get('analysis_type')
        # 1️⃣ Get custom prompt from form
        custom_prompt = request.form.get('custom_prompt', '').strip()

        # 2️⃣ Fallback to main prompt if custom_prompt is empty
        if not custom_prompt:
            print("⚠️ custom_prompt empty, using main PROMPTS fallback")
            # SERVICE_PROMPTS['Cargo_BL'] contains the default prompts per language
            svc_prompts_code = SERVICE_PROMPTS.get('Cargo_BL', {})
            # Pick English ('en') or any other language you want
            custom_prompt = svc_prompts_code.get('en') or ""

        keep_filename = request.form.get('keep_filename') == 'on'
        user = get_current_user()
        # 檢查使用者是否有該分析器權限
        allowed = get_user_allowed_analyzers(user['id'])
        if analysis_type not in allowed:
            return jsonify({'success': False, 'error': 'No permission to use this analyzer'}), 403

        # 解析用戶設定（避免在背景執行緒中取用 session）
        setting = get_user_analyzer_setting(user['id'], analysis_type)
        vlm_provider = setting.get('vlm_provider', 'cloud')
        vlm_model = setting.get('ollama_model')
        lang = setting.get('ocr_lang', 'auto')  # 從用戶設定獲取語言
        save_files = setting.get('save_files', False)  # 從用戶設定獲取是否保存文件
        username = session.get('username')
        
        if not files or all(f.filename == '' for f in files):
            return jsonify({'success': False, 'error': 'Please select files'})
        
        # 預處理檔案 - 先讀取所有檔案內容避免並行處理中的檔案關閉問題
        file_data_list = []
        for file in files:
            if file.filename != '':
                try:
                    file.seek(0)
                    file_content = file.read()
                    file_data_list.append({
                        'filename': file.filename,
                        'content': file_content
                    })
                except Exception as e:
                    app.logger.error(f"Failed to read file {file.filename}: {str(e)}")
                    continue
        
        if not file_data_list:
            return jsonify({'success': False, 'error': 'Unable to read any files'})
        
        # Initialize progress
        total_files = len(file_data_list)
        update_progress(task_id, 0, total_files, 'Starting analysis...')
        
        # Run analysis in background
        thread = threading.Thread(
            target=process_files_with_progress, 
            args=(task_id, file_data_list, analysis_type, lang, custom_prompt, keep_filename, vlm_provider, vlm_model, username, save_files)
        )
        thread.start()
        
        return jsonify({'success': True, 'task_id': task_id})
    
    except Exception as e:
        app.logger.error(f"Analysis error: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)})

def extract_worker(args):
    filename, file_path = args
    local_total_pages = 0
    content = None

    # Debug logs to check file existence
    print(f"[Worker] Processing file: {filename}")
    print(f"[Worker] File path: {file_path}")
    if not os.path.exists(file_path):
        print(f"[Worker][ERROR] File does not exist: {file_path}")
        return (filename, None, file_path, 0)

    try:
        if filename.lower().endswith('.pdf'):
            with pdfplumber.open(file_path) as pdf:
                page_count = len(pdf.pages)
                local_total_pages = page_count
                print(f"[Worker] PDF {filename} page count: {page_count}")
        elif filename.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp')):
            local_total_pages = 1
        else:
            print(f"[Worker][WARNING] Unsupported file type: {filename}")
    except Exception as e:
        print(f"[Worker][ERROR] Failed to open file {filename}: {e}")
        return (filename, None, file_path, 0)

    # Extract content using your existing extract_file_content function
    try:
        from app import extract_file_content  # ensure multi-process import works
        content = extract_file_content(file_path)
        if not content:
            print(f"[Worker][WARNING] Content is empty for file: {filename}")
    except Exception as e:
        print(f"[Worker][ERROR] extract_file_content failed for {filename}: {e}")
        content = None

    return (filename, content, file_path, local_total_pages)


def process_files_with_progress(task_id, file_data_list, analysis_type, lang, custom_prompt, keep_filename=True,
                                vlm_provider: Optional[str] = None, vlm_model: Optional[str] = None,
                                username: Optional[str] = None, save_files: bool = False):
    import pdfplumber
    import time
    try:
        saved_files = []
        start_time = time.time()
        total_files = len(file_data_list)
        processed_files = 0
        all_data = []
        total_pages = 0
        total_tokens = 0
        log_entries = []
        log_entries.append(f"=== File Analysis Processing Log ===")
        log_entries.append(f"Start Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        log_entries.append(f"Analysis Type: {analysis_type}")
        log_entries.append(f"Language: {lang}")
        log_entries.append(f"File Count: {total_files}")
        log_entries.append(f"Custom Prompt: {'Yes' if custom_prompt else 'No'}")
        log_entries.append("")

        update_progress(task_id, 0, 100, 'Starting analysis...')

        file_path_list = []
        for file_data in file_data_list:
            filename = file_data['filename']
            file_content = file_data['content']
            unique_filename = f"{uuid.uuid4().hex}_{secure_filename(filename)}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            with open(file_path, 'wb') as f:
                f.write(file_content)
                f.flush()
                os.fsync(f.fileno())
            try:
                os.chmod(file_path, 0o644)
            except Exception as e:
                app.logger.warning(f"Failed to set file permissions for {file_path}: {e}")
            file_path_list.append((filename, file_path))
            log_entries.append(f"File Preparation: {filename} ({len(file_content)} bytes)")

        log_entries.append("")
        log_entries.append("=== OCR Text Extraction Phase ===")

        extracted = []
        with ProcessPoolExecutor(max_workers=4) as executor:
            for idx, (filename, content, file_path, local_total_pages) in enumerate(
                    executor.map(extract_worker, file_path_list)):
                processed_files += 1
                progress_pct = int(10 + (processed_files / total_files * 40))  # 10-50%
                update_progress(task_id, progress_pct, 100, f"Extracting {filename}")
                extracted.append((filename, content, file_path))
                total_pages += local_total_pages
                content_length = len(content) if content else 0
                log_entries.append(f"OCR Extraction: {filename}")
                log_entries.append(f"  - Pages: {local_total_pages}")
                log_entries.append(f"  - Extracted Text Length: {content_length} characters")
                log_entries.append(f"  - Status: {'Success' if content_length > 0 else 'Failed'}")

        log_entries.append("")
        log_entries.append("=== LLM Analysis Phase ===")

        def extract_worker_multi_bl(file_path_tuple):
            """Modified extract worker for multi-BL PDFs"""
            filename, file_path = file_path_tuple

            # Check if PDF has multiple BLs
            if filename.lower().endswith('.pdf'):
                with pdfplumber.open(file_path) as pdf:
                    total_pages = len(pdf.pages)

                    # Check for multiple BLs markers
                    bl_markers = ['FIRST ORIGINAL', 'SECOND ORIGINAL', 'THIRD ORIGINAL']
                    bl_count = 0

                    full_text = ""
                    for page in pdf.pages:
                        text = page.extract_text() or ""
                        full_text += text + "\n"
                        for marker in bl_markers:
                            if marker in text:
                                bl_count += 1

                    # If multiple BLs detected, process differently
                    if bl_count > 4:  # More than one set of BLs
                        return process_multi_bl_separately(file_path, filename)
                    else:
                        # Single BL - use existing logic
                        return filename, full_text, file_path, total_pages

            # For non-PDF files, use existing logic
            return extract_worker(file_path_tuple)

        def process_multi_bl_separately(file_path, filename):
            """Process each BL in the PDF separately"""
            all_bl_texts = []

            with pdfplumber.open(file_path) as pdf:
                current_bl = ""
                bl_markers = ['FIRST ORIGINAL', 'SECOND ORIGINAL', 'THIRD ORIGINAL', 'COPY NOT NEGOTIABLE']

                for page in pdf.pages:
                    text = page.extract_text() or ""

                    # Check if new BL starts
                    is_new_bl = any(marker in text and 'Bill of Lading' in text for marker in bl_markers)

                    if is_new_bl and current_bl:
                        # Process previous BL
                        all_bl_texts.append(current_bl)
                        current_bl = text
                    else:
                        current_bl += "\n" + text

                # Add last BL
                if current_bl:
                    all_bl_texts.append(current_bl)

            # Combine all BLs with clear separation
            combined_text = "\n\n===NEXT_BL===\n\n".join(all_bl_texts)
            return filename, combined_text, file_path, len(all_bl_texts)

        def create_llm_worker(keep_filename_param):
            def llm_worker(args):
                filename, content, file_path = args

                if not content:
                    print("⚠️ Alert: content is None or empty!")
                    return [], 0

                total_tokens = 0
                all_results = []

                # Split content into sections
                bl_sections = content.split("===NEXT_BL===") if "===NEXT_BL===" in content else [content]

                # Check if all sections are empty
                if not any(section.strip() for section in bl_sections):
                    print("⚠️ Alert: all sections are empty!")
                    return [], 0

                print(f"DEBUG → Number of BL sections: {len(bl_sections)}")
                print("DEBUG → Sections preview (first 200 chars each):")
                for idx, section in enumerate(bl_sections, 1):
                    print(f"Section {idx}: {section[:200]!r}")

                # Helper function to clean VOY.NO.
                def clean_voy_no(voy_value, bl_number=None):
                    """Extract 4-digit voyage number, with fallback to BL number"""

                    # First try to clean the provided value
                    if voy_value and voy_value not in ['Not Available', 'null', None, '']:
                        # Check if value indicates "not found"
                        if not any(phrase in str(voy_value).lower() for phrase in
                                   ['not found', 'not available', 'n/a', '找不到']):
                            # Extract digits
                            cleaned = ''.join(filter(str.isdigit, str(voy_value)))
                            # Must be exactly 4 digits
                            if len(cleaned) == 4:
                                return cleaned

                    # Fallback: Extract from BL number (YYMM format)
                    if bl_number:
                        # BL format: "EMA 2406 0601" or "FLO 2505 0114"
                        parts = str(bl_number).strip().split()
                        if len(parts) >= 2:
                            # Second part should be YYMM (4 digits)
                            potential_voy = ''.join(filter(str.isdigit, parts[1]))
                            if len(potential_voy) == 4:
                                logging.info(f"📌 Extracted VOY.NO. from BL number '{bl_number}': {potential_voy}")
                                return potential_voy

                    return 'Not Available'

                # Helper function to clean record
                def clean_record(record):
                    """Clean VOY.NO. field in a record"""
                    if isinstance(record, dict) and 'VOY.NO.' in record:
                        original_voy = record['VOY.NO.']
                        bl_number = record.get('BL number')
                        cleaned_voy = clean_voy_no(original_voy, bl_number)
                        if original_voy != cleaned_voy:
                            print(f"🧹 Cleaned VOY.NO.: '{original_voy}' → '{cleaned_voy}' (BL: {bl_number})")
                        record['VOY.NO.'] = cleaned_voy
                    return record

                # Process each section
                for i, bl_section in enumerate(bl_sections, 1):
                    print(f"\n=== DEBUG: Processing BL Section {i}/{len(bl_sections)} ===")
                    print(f"Section length: {len(bl_section)} characters")
                    print(f"First 400 chars of section:\n{bl_section[:400]!r}")
                    print(f"analysis_type: {analysis_type}")
                    print(f"lang: {lang}")
                    if custom_prompt:
                        print(f"🧠 custom_prompt (first 8 words): {' '.join(custom_prompt.split()[:8])}")

                    try:
                        result, tokens_used = analyze_text_with_prompt_with_gemini(
                            bl_section, analysis_type, lang, custom_prompt
                        )
                        # Add this logging
                        print(f"DEBUG → LLM extracted for {filename}:")
                        if isinstance(result, list):
                            for item in result:
                                if isinstance(item, dict):
                                    bl_num = item.get('BL number', 'N/A')
                                    bl_qty = item.get('B/L quantity (MT)', 'N/A')
                                    split_qty = item.get('B/L split quantity (MT)', 'N/A')
                                    print(f"  BL: {bl_num} | B/L Qty: {bl_qty} | Split: {split_qty}")
                        total_tokens += tokens_used

                        print(f"DEBUG → Tokens used: {tokens_used}")
                        print("DEBUG → Raw result type:", type(result))
                        print("DEBUG → Raw result value:", repr(result))

                        if isinstance(result, dict):
                            print("DEBUG → Result keys:", list(result.keys()))

                    except Exception as e:
                        print(f"⚠️ ERROR while analyzing section {i}: {e}")
                        continue

                    # Skip if result is None or has error
                    if not result or (isinstance(result, dict) and result.get('error')):
                        print(f"⚠️ Skipping section {i} — result is None or contains error.")
                        continue

                    print(f"✅ Section {i} processed successfully.\n")

                    # Clean VOY.NO. in results
                    if isinstance(result, list):
                        result = [clean_record(item) for item in result]
                    elif isinstance(result, dict):
                        result = clean_record(result)

                    # Attach filename if requested
                    if keep_filename_param:
                        if isinstance(result, dict):
                            result['__filename__'] = filename
                        elif isinstance(result, list):
                            for item in result:
                                if isinstance(item, dict):
                                    item['__filename__'] = filename

                    # Flatten results into a single list
                    if isinstance(result, list):
                        all_results.extend(result)
                    else:
                        all_results.append(result)

                return all_results, total_tokens

            return llm_worker

        llm_worker = create_llm_worker(keep_filename)

        processed_files_llm = 0
        file_processing_logs = {}

        with ThreadPoolExecutor(max_workers=4) as executor:
            for idx, (result, tokens_used) in enumerate(executor.map(llm_worker, extracted)):
                processed_files_llm += 1
                filename = extracted[idx][0]
                file_path = extracted[idx][2]
                file_processing_logs[filename] = {
                    'llm_tokens': tokens_used,
                    'vlm_fields': [],
                    'vlm_tokens': 0,
                    'extracted_fields': 0,
                    'missing_fields': 0
                }
                log_entries.append(f"LLM Analysis: {filename}")
                log_entries.append(f"  - Tokens Used: {tokens_used}")
                log_entries.append(f"  - Analysis Result: {'Success' if result else 'Failed'}")
                progress_pct = int(50 + (processed_files_llm / total_files * 40))  # 50-90%
                update_progress(task_id, progress_pct, 100, f"Analyzing {filename}")

                if result:
                    if analysis_type == 'Cargo_BL':

                        def validate_bl_data(item):
                            """Validate B/L vs split quantity with detailed logging."""
                            bl_qty = item.get('B/L quantity (MT)')
                            split_qty = item.get('B/L split quantity (MT)')
                            bl_number = item.get('BL number', 'Unknown')

                            # Debug log
                            app.logger.info(f"🔍 Validating BL {bl_number}: B/L qty='{bl_qty}', Split qty='{split_qty}'")

                            try:
                                # Clean and convert B/L quantity
                                if bl_qty not in [None, '', 'n/a', 'null']:
                                    bl_val = float(str(bl_qty).replace(',', '').strip())
                                else:
                                    bl_val = None

                                # Clean and convert split quantity
                                if split_qty not in [None, '', 'n/a', 'null']:
                                    split_val = float(str(split_qty).replace(',', '').strip())
                                else:
                                    split_val = None

                                app.logger.info(f"🔍 Converted: bl_val={bl_val}, split_val={split_val}")

                                # Validation logic
                                if split_val is not None and bl_val is not None:
                                    # Check if split > B/L (invalid)
                                    if split_val > bl_val + 0.001:
                                        app.logger.warning(
                                            f"❌ INVALID: BL {bl_number} - Split ({split_val}) > B/L ({bl_val}). Clearing."
                                        )
                                        item['B/L split quantity (MT)'] = None

                                    # Check if split == B/L (suspicious - likely error)
                                    elif abs(split_val - bl_val) < 0.001:
                                        app.logger.warning(
                                            f"⚠️ EQUALS: BL {bl_number} - Split ({split_val}) == B/L ({bl_val}). Clearing."
                                        )
                                        item['B/L split quantity (MT)'] = None

                                    # Valid split quantity
                                    else:
                                        app.logger.info(f"✅ Valid: BL {bl_number} - {split_val} <= {bl_val}")
                                        item['B/L split quantity (MT)'] = f"{split_val:,.3f}"

                                # Format quantities
                                if split_val is None or item.get('B/L split quantity (MT)') is None:
                                    item['B/L split quantity (MT)'] = None

                                if bl_val is not None:
                                    item['B/L quantity (MT)'] = f"{bl_val:,.3f}"
                                else:
                                    item['B/L quantity (MT)'] = None

                            except Exception as e:
                                app.logger.error(f"❌ Error validating BL {bl_number}: {e}")

                            return item

                        def fill_missing_fields(d):
                            nonlocal total_tokens
                            vlm_processed_fields = []
                            vlm_tokens_used = 0

                            # Log the BL number we're processing
                            bl_number = d.get('BL number', 'Unknown')
                            app.logger.info(f"🔍 Processing BL: {bl_number}")

                            extracted_count = sum(1 for v in d.values() if not is_missing(v))
                            missing_count = sum(1 for v in d.values() if is_missing(v))
                            file_processing_logs[filename]['extracted_fields'] = extracted_count
                            file_processing_logs[filename]['missing_fields'] = missing_count

                            for key, value in d.items():
                                # Skip copy_type field
                                if key == 'copy_type':
                                    continue

                                # Log what we found for this field
                                if not is_missing(value):
                                    app.logger.info(f"  ✓ {key}: '{value}' (already extracted)")
                                else:
                                    app.logger.info(f"  ✗ {key}: Missing, will use VLM")

                                if is_missing(value):
                                    # Pass B/L quantity if we're filling split quantity
                                    bl_qty = d.get('B/L quantity (MT)') if key == 'B/L split quantity (MT)' else None

                                    # Log before VLM
                                    if bl_qty:
                                        app.logger.info(f"    → VLM will validate against B/L qty: {bl_qty}")

                                    filled, vlm_tokens = gemini_vlm_field(
                                        file_path,
                                        key,
                                        lang,
                                        provider_override=vlm_provider,
                                        model_override=vlm_model,
                                        bl_quantity=bl_qty
                                    )

                                    # Log after VLM
                                    app.logger.info(f"    → VLM returned: '{filled}'")

                                    d[key] = filled
                                    total_tokens += vlm_tokens
                                    vlm_tokens_used += vlm_tokens
                                    vlm_processed_fields.append(f"{key}={filled}")
                                    logging.info(
                                        f"VLM 補救: 檔案={os.path.basename(file_path)}, 欄位={key}, 補救值={filled}")

                            file_processing_logs[filename]['vlm_fields'] = vlm_processed_fields
                            file_processing_logs[filename]['vlm_tokens'] = vlm_tokens_used
                            return d

                        def fill_missing_charterers_by_cargo(data_list):
                            """Fill missing charterers based on same Cargo # having same charterer"""

                            # Group by Cargo #
                            cargo_groups = {}
                            for item in data_list:
                                cargo_num = item.get('Cargo #')
                                if cargo_num:
                                    if cargo_num not in cargo_groups:
                                        cargo_groups[cargo_num] = []
                                    cargo_groups[cargo_num].append(item)

                            # For each cargo group, find the charterer and apply to all
                            for cargo_num, items in cargo_groups.items():
                                # Find first valid charterer in this group
                                valid_charterer = None
                                for item in items:
                                    charterer = item.get('Charterer')
                                    if charterer and charterer not in ['Not Available', 'null', None, '']:
                                        valid_charterer = charterer
                                        break

                                # Apply to all items in this cargo group
                                if valid_charterer:
                                    for item in items:
                                        current_charterer = item.get('Charterer')
                                        if not current_charterer or current_charterer == 'Not Available':
                                            item['Charterer'] = valid_charterer
                                            app.logger.info(
                                                f"🔄 Filled Charterer for Cargo #{cargo_num}, BL {item.get('BL number')}: '{valid_charterer}'")

                            return data_list

                    if isinstance(result, dict) and 'data' in result:
                        cargo_data = result.get('data', [])
                        if isinstance(cargo_data, list):
                            for d in cargo_data:
                                fill_missing_fields(d)
                                validate_bl_data(d)
                                all_data.append(d)
                    elif isinstance(result, list):
                        for d in result:
                            fill_missing_fields(d)
                            validate_bl_data(d)
                            all_data.append(d)
                    elif isinstance(result, dict):
                        fill_missing_fields(result)
                        validate_bl_data(result)
                        all_data.append(result)
                    else:
                        if isinstance(result, dict):
                            all_data.append(result)
                        elif isinstance(result, list):
                            all_data.extend(result)

                    total_tokens += tokens_used

                    # After processing all files, fill missing charterers by cargo group
                    all_data = fill_missing_charterers_by_cargo(all_data)

        # Postprocessing
        update_progress(task_id, 90, 100, 'Post-processing data...')

        import re
        import pprint
        def clean_and_convert_qty(value):
            """
            Robustly cleans quantity-related strings by removing common units,
            handles both European (5.349.878) and US (5,349.878) formats.
            """
            if value is None or value in ['', 'null', 'n/a']:
                return None

            # Convert to string
            text = str(value)

            # Remove common units
            text = text.replace('MTS', '').replace('MT', '').replace('tons', '').replace('噸', '').strip()

            # Count dots and commas to determine format
            dot_count = text.count('.')
            comma_count = text.count(',')

            # European format with multiple dots (5.349.878)
            if dot_count > 1:
                # Remove thousand separators, keep last dot as decimal
                parts = text.split('.')
                if len(parts) > 1:
                    text = ''.join(parts[:-1]) + '.' + parts[-1]
            # US format with commas (5,349.878 or 5,349,878)
            elif comma_count >= 1:
                text = text.replace(',', '')

            # Extract the number
            match = re.search(r'([\d\.]+)', text)
            if match:
                try:
                    return float(match.group(1))
                except (ValueError, TypeError):
                    return None
            return None

        if all_data and analysis_type == 'Cargo_BL':
            log_entries = []  # Ensure log_entries is initialized

            # --- SINGLE, UNIFIED PROCESSING LOOP ---
            # All data cleaning and validation happens here, once per item.
            for item in all_data:
                # 1. Apply OCR fixes to all string values first
                for key, value in item.items():
                    if isinstance(value, str):
                        item[key] = fix_ocr_errors(value)

                # 2. Clean and convert B/L and Split quantities
                bl_val = clean_and_convert_qty(item.get('B/L quantity (MT)'))
                split_val = clean_and_convert_qty(item.get('B/L split quantity (MT)'))

                # 3. Apply validation logic for Split Quantity
                if split_val is not None and bl_val is not None:
                    # Use a small tolerance for float comparison
                    if split_val > bl_val + 0.001:
                        # ❌ Split > B/L → Invalid, set to None
                        item['B/L split quantity (MT)'] = None
                        log_entries.append(
                            f"SID [WARN] Split qty > B/L qty for BL={item.get('BL number', 'N/A')} | "
                            f"Split={split_val} > B/L={bl_val} → cleared."
                        )
                    elif abs(split_val - bl_val) < 0.001:
                        # ⚠️ Split == B/L → Suspicious (likely no actual split), set to None
                        item['B/L split quantity (MT)'] = None
                        log_entries.append(
                            f"SID [WARN] Split qty equals B/L qty for BL={item.get('BL number', 'N/A')} | "
                            f"Split={split_val} == B/L={bl_val} → cleared (no actual split)."
                        )
                    else:
                        # ✅ Valid split: split < B/L
                        item['B/L split quantity (MT)'] = f"{split_val:,.3f}"
                        log_entries.append(
                            f"SID [INFO] Valid split qty for BL={item.get('BL number', 'N/A')} | "
                            f"Split={split_val} < B/L={bl_val} → kept."
                        )
                elif split_val is not None:
                    # Split exists but no B/L quantity to compare against (unusual)
                    item['B/L split quantity (MT)'] = f"{split_val:,.3f}"
                else:
                    # No split quantity found
                    item['B/L split quantity (MT)'] = None

                # 4. Format B/L quantity with commas
                if bl_val is not None:
                    item['B/L quantity (MT)'] = f"{bl_val:,.3f}"
                else:
                    item['B/L quantity (MT)'] = None

                # 5. Validate OBL release date format
                obl_date = item.get('OBL release date')
                if not (isinstance(obl_date, str) and re.match(r'^\d{4}-\d{2}-\d{2}$', obl_date)):
                    item['OBL release date'] = None

            app.logger.info("=== DEBUG: all_data BEFORE cargo_bl_postprocess ===")
            if all_data:
                df = pd.DataFrame(all_data)
                app.logger.info("\n" + tabulate(df, headers='keys', tablefmt='psql', showindex=True))
            else:
                app.logger.info("⚠️ all_data is empty before cargo_bl_postprocess.")

            # ALWAYS call cargo_bl_postprocess (moved outside the if/else)
            from analyzers.services import cargo_bl_postprocess
            all_data = cargo_bl_postprocess(all_data, keep_filename)

            # Debug: Check after calling cargo_bl_postprocess
            has_filename_after_postprocess = any('File name' in item for item in all_data if isinstance(item, dict))
            if has_filename_after_postprocess:
                sample_item = next((item for item in all_data if isinstance(item, dict) and 'File name' in item), None)
                if sample_item:
                    app.logger.info(f"Sample item with 'File name': {sample_item.get('File name')}")

            # 獲取列名（cargo_bl_postprocess 已經處理了 __filename__ 字段）
            headers = list(all_data[0].keys()) if all_data else []

            # 調試日誌
            app.logger.info(f"keep_filename: {keep_filename}")
            app.logger.info(f"Final headers: {headers}")
            app.logger.info(f"File name in headers: {'File name' in headers}")

            rows = []
            for i, item in enumerate(all_data):
                row = []
                for header in headers:
                    value = item.get(header, '')
                    row.append(str(value) if value is not None else '')
                is_total_row = i == len(all_data) - 2
                is_cgos_row = i == len(all_data) - 1
                rows.append({
                    'data': row,
                    'is_total': is_total_row,
                    'is_cgos': is_cgos_row
                })
            excel_filename = f"{analysis_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            # 獲取當前用戶信息
            current_user = get_user_by_username(username)
            user_info = {
                'display_name': current_user.get('display_name') if current_user else None,
                'logo_file': current_user.get('logo_file') if current_user else None
            }

            # 檢查是否有 VESSEL 和 VOY.NO. 信息（針對 Cargo_BL）
            vessel_voy_info = globals().get('vessel_voy_data', None)

            app.logger.info("=== DEBUG: all_data BEFORE export_to_excel ===")
            if all_data:
                # Remove copy_type column if it still exists
                for item in all_data:
                    if 'copy_type' in item:
                        del item['copy_type']

                df = pd.DataFrame(all_data)
                app.logger.info("\n" + tabulate(df, headers='keys', tablefmt='psql', showindex=True))
            else:
                app.logger.info("⚠️ all_data is empty.")

            update_progress(task_id, 95, 100, 'Generating Excel file...')

            excel_path = export_to_excel(all_data, excel_filename, 'Hansa Tankers', user_info=user_info,
                                         include_logo=True, vessel_voy_info=vessel_voy_info,
                                         keep_filename=keep_filename)
            # 生成對應的 log 檔案，保存在 logs 資料夾中
            log_filename = excel_filename.replace('.xlsx', '_log.txt')

            # 確保 logs 資料夾存在
            logs_dir = 'logs'
            if not os.path.exists(logs_dir):
                os.makedirs(logs_dir)

            # Define log path and time cost
            log_path = os.path.join(logs_dir, log_filename)
            time_cost = time.time() - start_time

            # === ADD MORE DETAILED SUMMARY BEFORE END TIME ===
            log_entries.append("")
            log_entries.append("=== Processing Summary ===")
            log_entries.append(f"Total Processing Time: {time_cost:.2f} seconds")
            log_entries.append(f"Total Tokens Used: {total_tokens}")
            log_entries.append(f"Total Pages: {total_pages}")
            log_entries.append(f"Successfully Processed Files: {len(extracted)}")
            log_entries.append("")

            # === ADD DETAILED FILE PROCESSING STATISTICS ===
            log_entries.append("=== Detailed File Processing Statistics ===")
            for fname, flog in file_processing_logs.items():
                log_entries.append(f"File: {fname}")
                log_entries.append(f"  - LLM Tokens: {flog['llm_tokens']}")
                log_entries.append(f"  - VLM Tokens: {flog['vlm_tokens']}")
                log_entries.append(f"  - Total Tokens: {flog['llm_tokens'] + flog['vlm_tokens']}")
                log_entries.append(f"  - Successfully Extracted Fields: {flog['extracted_fields']}")
                log_entries.append(f"  - Missing Fields: {flog['missing_fields']}")
                if flog['vlm_fields']:
                    log_entries.append(f"  - VLM Remediated Fields: {', '.join(flog['vlm_fields'])}")
                log_entries.append("")

            # === FINAL SUMMARY ===
            log_entries.append(f"End Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            log_entries.append(f"Output Excel: {excel_filename}")
            log_entries.append(f"Output Log: {log_filename}")

            # === WRITE LOG FILE ===
            with open(log_path, 'w', encoding='utf-8') as log_file:
                log_file.write('\n'.join(log_entries))

            logging.info(f"Detailed log file generated: {log_filename}")

            time_cost = round(time.time() - start_time, 2)

            # 根據設置決定是否保存文件
            saved_files = []
            if save_files:
                # 保存文件，記錄文件路徑到歷史記錄中
                for filename, _, file_path in extracted:
                    if os.path.exists(file_path):
                        # 重命名文件為更友好的名稱
                        file_ext = os.path.splitext(filename)[1]
                        new_filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{secure_filename(filename)}"
                        new_file_path = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)
                        try:
                            os.rename(file_path, new_file_path)
                            saved_files.append((filename, new_filename))
                            log_entries.append(f"File Saved: {filename} -> {new_filename}")
                        except Exception as e:
                            app.logger.warning(f"Failed to rename saved file: {file_path} {e}")
                            # 如果重命名失敗，保持原文件名
                            saved_files.append((filename, os.path.basename(file_path)))
            else:
                # 統一刪除所有臨時檔案
                for _, _, file_path in extracted:
                    try:
                        if os.path.exists(file_path):
                            os.remove(file_path)
                    except Exception as e:
                        app.logger.warning(f"Failed to delete temporary file: {file_path} {e}")

            history_entry = {
                'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'files': [f['filename'] for f in file_data_list],
                'analysis_type': analysis_type,
                'lang': lang,
                'rows': len(all_data),
                'cols': len(headers),
                'headers': headers,
                'tokens': total_tokens,
                'excel': excel_filename,
                'log_file': log_filename,
                'time_cost': time_cost,
                'duration_str': f"{time_cost:.2f} sec.",
                'username': username,
                'total_pages': total_pages,
                'saved_files': saved_files
            }
            save_to_history(history_entry)

            app.logger.info(f"About to set final progress for task {task_id}")
            update_progress(task_id, 100, 100, 'Analysis completed!')
            app.logger.info(f"Updated final progress for task {task_id}")

            # 提取純數據用於前端顯示
            simple_rows = [row['data'] if isinstance(row, dict) and 'data' in row else row for row in rows]
            app.logger.info(f"About to set final_result for task {task_id}: prepared {len(simple_rows)} rows")

            with progress_lock:
                progress_store[task_id]['final_result'] = {
                    'success': True,
                    'headers': headers,
                    'rows': simple_rows,
                    'excel': excel_filename,
                    'log_file': log_filename,
                    'lang_code': lang,
                    'total_pages': total_pages
                }
                app.logger.info(
                    f"Successfully set final_result for task {task_id}: success=True, rows={len(simple_rows)}, excel={excel_filename}")
        else:
            app.logger.warning(f"No data extracted for task {task_id}, setting failure result")
            with progress_lock:
                progress_store[task_id]['final_result'] = {
                    'success': False,
                    'error': 'Unable to extract valid data from files'
                }
                app.logger.info(f"Set final_result for task {task_id}: success=False (no data)")
        # 將保存的文件信息添加到進度存儲中
        if saved_files:
            with progress_lock:
                progress_store[task_id]['saved_files'] = saved_files
    except Exception as e:
        app.logger.error(f"Error occurred during processing: {str(e)}", exc_info=True)
        with progress_lock:
            progress_store[task_id]['final_result'] = {
                'success': False,
                'error': str(e)
            }
            app.logger.info(f"Set final_result for task {task_id}: success=False due to exception: {str(e)}")

@app.route('/progress/<task_id>')
def check_progress(task_id):
    """Check task progress"""
    return get_progress(task_id)

@app.route('/export/<filename>')
def export(filename):
    # First check uploads directory
    upload_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(upload_path):
        return send_file(upload_path, as_attachment=True, download_name=filename)
    
    # Then check temp directory
    temp_path = os.path.join(tempfile.gettempdir(), filename)
    if os.path.exists(temp_path):
        return send_file(temp_path, as_attachment=True, download_name=filename)
    
    return "File not found", 404

@app.route('/logs/<filename>')
def download_log(filename):
    """Download log file"""
    # Check logs directory
    logs_path = os.path.join('logs', filename)
    if os.path.exists(logs_path):
        return send_file(logs_path, as_attachment=True, download_name=filename)
    
    return "Log file not found", 404

@app.route('/saved_files/<filename>')
def download_saved_file(filename):
    """Download saved analysis file"""
    # Check uploads directory for saved files
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True, download_name=filename)
    
    return "File not found", 404




@app.route('/download_zip/<int:index>')
def download_zip_files(index):
    """Download all saved files as ZIP for a specific analysis record"""
    try:
        # Align index with admin dashboard ordering (time desc)
        def _get_history_time(x):
            return x.get('time') or x.get('timestamp') or ''
        sorted_history = sorted(history, key=_get_history_time, reverse=True)

        # Get the history record by index (from sorted list)
        if index < 0 or index >= len(sorted_history):
            app.logger.error(f"History index {index} not found, total history: {len(sorted_history)} (sorted)")
            return "Analysis record not found", 404

        record = sorted_history[index]
        saved_files = record.get('saved_files', [])
        
        app.logger.info(f"ZIP download request for index {index}, saved_files: {saved_files}")
        
        if not saved_files:
            app.logger.warning(f"No saved files for index {index}")
            return "No saved files found for this analysis record", 404
        
        # Create ZIP file in uploads directory (same as other files)
        import zipfile
        from datetime import datetime
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        zip_filename = f"analysis_files_{index}_{timestamp}.zip"
        zip_path = os.path.join(app.config['UPLOAD_FOLDER'], zip_filename)
        
        app.logger.info(f"Creating ZIP file at: {zip_path}")
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            files_added = 0
            for original_name, saved_name in saved_files:
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], saved_name)
                app.logger.info(f"Checking file: {file_path}")
                
                if os.path.exists(file_path):
                    # Add file to ZIP with original name
                    zipf.write(file_path, original_name)
                    files_added += 1
                    app.logger.info(f"Added {original_name} to ZIP from {file_path}")
                else:
                    app.logger.warning(f"File not found: {file_path}")
        
        if files_added == 0:
            if os.path.exists(zip_path):
                os.remove(zip_path)
            app.logger.error("No files could be added to ZIP")
            return "No files could be added to ZIP", 404
        
        app.logger.info(f"ZIP file created successfully with {files_added} files")
        
        # Use the same route as other downloads
        return send_file(
            zip_path,
            as_attachment=True,
            download_name=f"analysis_files_{record.get('time', 'unknown').replace(':', '-').replace(' ', '_')}.zip"
        )
        
    except Exception as e:
        app.logger.error(f"Error creating ZIP file: {e}", exc_info=True)
        return f"Error creating ZIP file: {str(e)}", 500


@app.route('/history')
@login_required
def get_history():
    import json
    username = session.get('username')

    conn = get_db_connection()
    rows = conn.execute('''
        SELECT * FROM history WHERE username = ? ORDER BY timestamp DESC
    ''', (username,)).fetchall()
    conn.close()

    history = []
    for row in rows:
        history.append({
            'time': row['timestamp'],
            'files': json.loads(row['files']),
            'analysis_type': row['analysis_type'],
            'lang': row['lang'],
            'rows': row['rows'],
            'cols': row['cols'],
            'headers': json.loads(row['headers']),
            'tokens': row['tokens'],
            'excel': row['excel'],
            'log_file': row['log_file'],
            'time_cost': row['time_cost'],
            'duration_str': row['duration_str'],
            'username': row['username'],
            'total_pages': row['total_pages'],
            'saved_files': json.loads(row['saved_files']) if row['saved_files'] else []
        })

    return render_template('history.html', history=history, config=config, lang='en')
@app.route('/clear_history', methods=['POST'])
def clear_history():
    # 先刪除所有已存的 Excel 檔案
    for h in history:
        excel_name = h.get('excel')
        if excel_name:
            excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_name)
            try:
                if os.path.exists(excel_path):
                    os.remove(excel_path)
            except Exception:
                pass
    history.clear()
    save_history()
    return jsonify({'ok': True})

def fix_ocr_errors(text):
    """Fix common OCR errors"""
    # Fix Row -> Stow
    text = re.sub(r'\bRow\b', 'Stow', text, flags=re.IGNORECASE)
    # Can add more OCR error correction rules
    return text


def enforce_column_order_list(data_list):
    """Enforce column order for list data"""
    if not data_list:
        return data_list
    
    desired_order = [
        'Cargo #', 'BL number', 'B/L quantity (MT)', 'B/L split quantity (MT)',
        'Cargo name', 'Charterer', 'Consignee, order to', 'Notify',
        'Stow', 'LoadPort', 'Disch. Port', 'OBL release date', 'Release cargo against'
    ]
    
    # Reorder each item's columns
    ordered_data = []
    for item in data_list:
        ordered_item = {}
        
        # First add columns in desired_order
        for col in desired_order:
            if col in item:
                ordered_item[col] = item[col]
        
        # Then add other columns
        for key, value in item.items():
            if key not in ordered_item:
                ordered_item[key] = value
        
        ordered_data.append(ordered_item)
    
    return ordered_data

def enforce_column_order(df, desired_order):
    """Reorder DataFrame columns according to specified order"""
    # Get existing columns
    existing_cols = df.columns.tolist()
    
    # Create new column order
    ordered_cols = []
    
    # First add columns that exist in desired_order
    for col in desired_order:
        if col in existing_cols:
            ordered_cols.append(col)
    
    # Then add other columns not in desired_order
    for col in existing_cols:
        if col not in ordered_cols:
            ordered_cols.append(col)
    
    return df[ordered_cols]

def sort_by_bl_number(df):
    """Sort by BL number"""
    if 'BL number' in df.columns:
        # Remove empty values and sort
        df_sorted = df.dropna(subset=['BL number']).sort_values('BL number')
        df_empty = df[df['BL number'].isna()]
        return pd.concat([df_sorted, df_empty], ignore_index=True)
    return df

def post_process_cargo_bl_data(df):
    """Post-process Cargo_BL data, handle special rules"""
    if 'Release cargo against' in df.columns:
        # Replace N/A with LOI
        df['Release cargo against'] = df['Release cargo against'].replace('N/A', 'LOI')
        df['Release cargo against'] = df['Release cargo against'].fillna('LOI')

    # Add total row
    if len(df) > 0:
        # 1. Total row: Calculate sum of B/L quantity and B/L split quantity
        total_row = {}
        for col in df.columns:
            if col == 'Cargo #':
                total_row[col] = 'Total'
            elif col == 'B/L quantity (MT)':
                # Calculate sum, handle possible string format (e.g., "1,007.659")
                total_sum = 0
                for val in df[col].dropna():
                    try:
                        # Remove commas and convert to float
                        clean_val = str(val).replace(',', '').strip()
                        if clean_val and clean_val != 'nan':
                            total_sum += float(clean_val)
                    except (ValueError, TypeError):
                        continue
                total_row[col] = f"{total_sum:,.3f}"
            elif col == 'B/L split quantity (MT)':
                # Calculate sum
                total_sum = 0
                for val in df[col].dropna():
                    try:
                        clean_val = str(val).replace(',', '').strip()
                        if clean_val and clean_val != 'nan':
                            total_sum += float(clean_val)
                    except (ValueError, TypeError):
                        continue
                total_row[col] = f"{total_sum:,.3f}"
            else:
                total_row[col] = ''

        # 2. # Cgos 行：計算 B/L number 的數量
        cgos_row = {}
        bl_count = df['BL number'].dropna().nunique() if 'BL number' in df.columns else 0

        for col in df.columns:
            if col == 'Cargo #':
                cgos_row[col] = f'# Cgos: {bl_count}'
            else:
                cgos_row[col] = ''

        # 將總計行添加到 DataFrame
        df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
        df = pd.concat([df, pd.DataFrame([cgos_row])], ignore_index=True)

    return df

def safe_float(val):
    try:
        if isinstance(val, str):
            val = val.replace(',', '')
        return float(val)
    except Exception:
        return 0.0

def process_same_category_bl(data_list):
    """
    處理同類 B/L 的 B/L quantity 和 Cargo# 重複問題
    根據 BL number 的第10-11位相同但第12-13位不同來判斷同類
    
    例如：
    - "EMA 2406 0801" 和 "EMA 2406 0802" 
    - 第10-11位都是 "08"，第12-13位分別是 "01" 和 "02"
    - 它們屬於同一類，只有第一個保留 B/L quantity 和 Cargo#
    """
    if not data_list:
        return data_list
    
    # 按 BL number 分組
    bl_groups = {}
    
    for i, item in enumerate(data_list):
        bl_number = item.get('BL number', '').strip()
        if not bl_number:
            continue
            
        # 提取 BL number 的第10-11位用於分組
        # 格式：EMA 2406 0804
        #       123456789012345
        # 第10-11位是第三部分的前兩位
        if len(bl_number) >= 13:
            try:
                # 直接取第10-11位字符
                group_key = bl_number[9:11]  # 第10-11位 (0-based index)
                
                if group_key not in bl_groups:
                    bl_groups[group_key] = []
                bl_groups[group_key].append((i, item))
            except (IndexError, ValueError):
                # 如果 BL number 格式不符合預期，單獨處理
                continue

    # 處理每個分組
    for group_key, group_items in bl_groups.items():
        print(f"\n🔹 Processing group: {group_key} (items: {len(group_items)})")

        if len(group_items) > 1:  # 只處理有多個項目的分組
            # 按 BL number 排序，確保處理順序一致
            group_items.sort(key=lambda x: x[1].get('BL number', ''))
            print("  ↳ Sorted items by BL number.")

            prev_bl_number = None  # 用於記錄前一個 BL number

            # 逐一處理項目
            for idx, (original_index, item) in enumerate(group_items):
                current_bl = item.get('BL number', '')
                print(f"    • Item {idx + 1} (original index {original_index}): BL={current_bl}")

                if idx == 0:
                    # 第一個項目保持不變
                    print("      → First item in group: keeping original values.")
                    prev_bl_number = current_bl
                    continue

                # 清空 Cargo #
                if 'Cargo #' in item:
                    item['Cargo #'] = None
                    print("      → Cleared Cargo #")

                # 根據 BL number 決定是否清空 B/L quantity
                if current_bl == prev_bl_number:
                    if 'B/L quantity (MT)' in item:
                        item['B/L quantity (MT)'] = None
                    print("      → Same BL as previous, cleared B/L quantity (MT).")
                else:
                    # 如果不同，保留原始數值
                    print("      → New BL number detected, kept B/L quantity (MT).")
                    prev_bl_number = current_bl

        else:
            print(f"  ⚪ Group has only one item — skipped.")
    return data_list

def process_single_file(file_data, analysis_type, prompt_template, keep_filename):
    """處理單一檔案的函數，用於並行處理"""
    file, filename = file_data
    try:
        suffix = filename.split('.')[-1].lower()
        
        # 檢查檔案大小
        file.seek(0, 2)
        file_size = file.tell()
        file.seek(0)
        
        if file_size == 0:
            print(f"錯誤：檔案大小為 0 - {filename}")
            return None, 0
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.'+suffix) as tmp:
            file.seek(0)
            file.save(tmp.name)
            
            saved_size = os.path.getsize(tmp.name)
            if saved_size == 0:
                file.seek(0)
                content = file.read()
                if len(content) > 0:
                    with open(tmp.name, 'wb') as f:
                        f.write(content)
                    saved_size = os.path.getsize(tmp.name)
                else:
                    return None, 0
            
            # 文字擷取
            text = extract_text(tmp.name, suffix)
            # OCR 錯誤修正
            text = fix_ocr_errors(text)
            
            # 清理臨時檔案
            try:
                os.unlink(tmp.name)
            except:
                pass
            
            if len(text.strip()) < 10:  # 過濾太短的文字
                print(f"警告：檔案 {filename} 解析出的文字太短，跳過")
                return None, 0
            
            # 準備 prompt
            if '{text}' not in prompt_template:
                prompt_str = f"{prompt_template}\n\n檔案內容：\n{text}"
            else:
                prompt_str = prompt_template.replace('{text}', text)
            
            # 呼叫 API
            result_json, tokens_used = call_gemini_api(prompt_str)
            data = extract_json_from_response(result_json)
            
            if keep_filename:
                if isinstance(data, dict):
                    data['__filename__'] = filename
                elif isinstance(data, list):
                    for d in data:
                        if isinstance(d, dict):
                            d['__filename__'] = filename
            
            print(f"✓ 完成處理：{filename} (Token: {tokens_used})")
            return data, tokens_used
            
    except Exception as e:
        print(f"檔案處理錯誤 {filename}：{e}")
        return None, 0

@app.route('/export_custom', methods=['POST'])
def export_custom():
    try:
        data = request.get_json()
        headers = data.get('headers')
        rows = data.get('rows')
        manual_data = data.get('manual_data', {'headers': [], 'rows': []})
        include_logo = data.get('include_logo', False)
        preview_title = data.get('preview_title', 'Hansa Tankers')
        
        if not headers or not rows:
            return jsonify({'success': False, 'error': '缺少資料'})
        
        import pandas as pd
        # 產生唯一檔名
        filename = f"CustomExport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # 獲取當前用戶信息  
        current_user = get_user_by_username(session.get('username'))
        user_info = {
            'display_name': current_user.get('display_name') if current_user else None,
            'logo_file': current_user.get('logo_file') if current_user else None
        }
        
        # 直接呼叫 export_to_excel，帶入所有新參數
        excel_path = export_to_excel(
            [dict(zip(headers, row)) for row in rows], 
            filename, 
            preview_title,
            manual_data=manual_data,
            include_logo=include_logo,
            user_info=user_info
        )
        
        if excel_path is None:
            return jsonify({'success': False, 'error': 'Failed to create Excel file'})
        
        # 檢查檔案是否真的存在
        if not os.path.exists(excel_path):
            return jsonify({'success': False, 'error': f'Excel file not found: {excel_path}'})
        
        # 注意：這裡不再儲存歷史記錄，因為這只是重新匯出已分析的資料
        return jsonify({'success': True, 'download_url': url_for('download_custom_excel', filename=filename)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/download_custom_excel/<filename>')
def download_custom_excel(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

def is_missing(value):
    return value is None or str(value).strip() == '' or str(value).lower() in ['null', 'n/a']


def gemini_vlm_field(file_path, field_name, lang='zh', provider_override: Optional[str] = None,
                     model_override: Optional[str] = None, bl_quantity: Optional[str] = None):
    # 根據呼叫方指定或使用者設定決定使用 cloud 或 local(VLM: ollama)
    provider = provider_override
    ollama_model = model_override
    if provider is None:
        user = get_current_user()
        if user:
            setting = get_user_analyzer_setting(user['id'], 'Cargo_BL')
            provider = setting.get('vlm_provider', 'cloud')
            ollama_model = ollama_model or setting.get('ollama_model')
        else:
            provider = 'cloud'

    # Helper function to clean VOY.NO.
    def clean_voy_no(voy_value):
        """Extract only 4-digit voyage numbers, return 'Not Available' if invalid"""
        if not voy_value:
            return 'Not Available'

        # Check if value already indicates "not found"
        if any(phrase in str(voy_value).lower() for phrase in ['not found', 'not available', 'n/a', 'null', '找不到']):
            return 'Not Available'

        # Extract digits
        cleaned = ''.join(filter(str.isdigit, str(voy_value)))

        # VOY.NO. must be exactly 4 digits (YYMM format like 2406, 2505)
        if len(cleaned) == 4:
            return cleaned
        else:
            # Invalid format (too short, too long, or empty)
            logging.info(
                f"⚠️ Invalid VOY.NO. format: '{voy_value}' (cleaned: '{cleaned}', length: {len(cleaned)}) → 'Not Available'")
            return 'Not Available'

    # Helper function to validate split quantity
    def validate_split_qty(split_value, bl_value):
        """Validate split quantity against B/L quantity"""
        if not split_value or not bl_value:
            return split_value

        try:
            # Clean the values
            split_str = str(split_value).replace(',', '').replace('MT', '').replace('MTS', '').strip()
            bl_str = str(bl_value).replace(',', '').replace('MT', '').replace('MTS', '').strip()

            split_float = float(split_str)
            bl_float = float(bl_str)

            # If split > BL, invalid
            if split_float > bl_float + 0.001:
                logging.warning(f"⚠️ VLM extracted invalid split qty: {split_value} > B/L: {bl_value} → returning None")
                return None

            # If split == BL, might be error (no actual split found)
            if abs(split_float - bl_float) < 0.001:
                logging.warning(f"⚠️ VLM extracted split qty equals B/L: {split_value} == {bl_value} → returning None")
                return None

            # Valid split quantity
            logging.info(f"✓ VLM extracted valid split qty: {split_value} <= {bl_value}")
            return split_value

        except Exception as e:
            logging.warning(f"Error validating split qty: {e}")
            return None

    if provider == 'local':
        try:
            from analyzers.clients.ollama_client import call_ollama_vlm
            text, tokens = call_ollama_vlm(file_path, field_name, lang, model=ollama_model)

            # Clean VOY.NO. if needed
            if field_name == 'VOY.NO.':
                original = text
                text = clean_voy_no(text)
                if original != text:
                    logging.info(f"🧹 VLM Cleaned VOY.NO.: '{original}' → '{text}'")

            # Validate split quantity if needed
            elif field_name == 'B/L split quantity (MT)' and bl_quantity:
                original = text
                text = validate_split_qty(text, bl_quantity)
                if original != text:
                    logging.info(f"🧹 VLM Validated split qty: '{original}' → '{text}'")

            return text, tokens
        except Exception as e:
            logging.warning(f"Ollama VLM 失敗，改用雲端：{e}")
            from analyzers.services import vlm_fill_field as _vlm
            text, tokens = _vlm(file_path, field_name, lang)

            # Clean VOY.NO. if needed
            if field_name == 'VOY.NO.':
                original = text
                text = clean_voy_no(text)
                if original != text:
                    logging.info(f"🧹 VLM Cleaned VOY.NO.: '{original}' → '{text}'")

            # Validate split quantity if needed
            elif field_name == 'B/L split quantity (MT)' and bl_quantity:
                original = text
                text = validate_split_qty(text, bl_quantity)
                if original != text:
                    logging.info(f"🧹 VLM Validated split qty: '{original}' → '{text}'")

            return text, tokens
    else:
        from analyzers.services import vlm_fill_field as _vlm
        text, tokens = _vlm(file_path, field_name, lang)

        # Clean VOY.NO. if needed
        if field_name == 'VOY.NO.':
            original = text
            text = clean_voy_no(text)
            if original != text:
                logging.info(f"🧹 VLM Cleaned VOY.NO.: '{original}' → '{text}'")

        # Validate split quantity if needed
        elif field_name == 'B/L split quantity (MT)' and bl_quantity:
            original = text
            text = validate_split_qty(text, bl_quantity)
            if original != text:
                logging.info(f"🧹 VLM Validated split qty: '{original}' → '{text}'")

        return text, tokens

# =========================
# 管理者後台：使用者與分析器設定、歷史
# =========================

@app.route('/admin')
@admin_required
def admin_dashboard():
    conn = get_db_connection()
    rows = conn.execute('SELECT id, username, display_name, is_admin, is_active, email, phone, address, logo_file, notes FROM users ORDER BY id ASC').fetchall()
    conn.close()
    users = [dict(r) for r in rows]
    
    # 重新加載最新的歷史記錄
    global history
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, 'rb') as f:
                history = pickle.load(f)
        except Exception as e:
            app.logger.error(f"Failed to reload history: {e}")
    
    # 傳遞歷史資料供統計使用，按時間降序排序
    def get_history_time(x):
        return x.get('time') or x.get('timestamp') or ''
    sorted_history = sorted(history, key=get_history_time, reverse=True)
    
    # Create user mapping table (username -> user info)
    user_map = {u['username']: u for u in users}
    
    # Calculate statistics
    stats = calculate_user_statistics(sorted_history)
    
    return render_template('admin_dashboard.html', users=users, history=sorted_history, user_map=user_map, selected_username='', stats=stats)

@app.route('/admin/users/add', methods=['POST'])
@admin_required
def admin_add_user():
    username = request.form.get('username', '').strip()
    display_name = request.form.get('display_name', '').strip()
    password = request.form.get('password', '').strip()
    email = request.form.get('email', '').strip()
    phone = request.form.get('phone', '').strip()
    address = request.form.get('address', '').strip()
    notes = request.form.get('notes', '').strip()
    is_admin = 1 if request.form.get('is_admin') == 'on' else 0
    
    if not username or not password:
        return redirect(url_for('admin_dashboard'))
    
    # 處理 logo 文件上傳
    logo_filename = None
    if 'logo' in request.files:
        logo_file = request.files['logo']
        if logo_file and logo_file.filename:
            # 確保 uploads 目錄存在
            import uuid
            import os
            from werkzeug.utils import secure_filename
            
            filename = secure_filename(logo_file.filename)
            file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
            if file_ext in ['jpg', 'jpeg', 'png', 'gif']:
                logo_filename = f"user_logo_{uuid.uuid4().hex[:8]}.{file_ext}"
                logo_path = os.path.join(app.config['UPLOAD_FOLDER'], logo_filename)
                logo_file.save(logo_path)
    
    try:
        conn = get_db_connection()
        conn.execute(
            '''INSERT INTO users (username, display_name, password_hash, is_admin, is_active, email, phone, address, logo_file, notes) 
               VALUES (?, ?, ?, ?, 1, ?, ?, ?, ?, ?)''',
            (username, display_name or None, generate_password_hash(password, method='pbkdf2:sha256'), is_admin, 
             email or None, phone or None, address or None, logo_filename, notes or None)
        )
        conn.commit()
    except Exception as e:
        logging.error(f"Failed to add user: {e}")
    finally:
        conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/users/edit', methods=['POST'])
@admin_required
def admin_edit_user():
    user_id = request.form.get('user_id')
    username = request.form.get('username', '').strip()
    display_name = request.form.get('display_name', '').strip()
    password = request.form.get('password', '').strip()
    email = request.form.get('email', '').strip()
    phone = request.form.get('phone', '').strip()
    address = request.form.get('address', '').strip()
    notes = request.form.get('notes', '').strip()
    is_admin = 1 if request.form.get('is_admin') == 'on' else 0
    
    if not user_id or not username:
        return redirect(url_for('admin_dashboard'))
    
    # 處理 logo 文件上傳
    logo_filename = None
    if 'logo' in request.files:
        logo_file = request.files['logo']
        if logo_file and logo_file.filename:
            import uuid
            import os
            from werkzeug.utils import secure_filename
            
            filename = secure_filename(logo_file.filename)
            file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
            if file_ext in ['jpg', 'jpeg', 'png', 'gif']:
                logo_filename = f"user_logo_{uuid.uuid4().hex[:8]}.{file_ext}"
                logo_path = os.path.join(app.config['UPLOAD_FOLDER'], logo_filename)
                logo_file.save(logo_path)
    
    try:
        conn = get_db_connection()
        
        # 如果有新密碼，則更新密碼
        if password:
            if logo_filename:
                conn.execute(
                    '''UPDATE users SET username=?, display_name=?, password_hash=?, is_admin=?, email=?, phone=?, address=?, logo_file=?, notes=? 
                       WHERE id=?''',
                    (username, display_name or None, generate_password_hash(password, method='pbkdf2:sha256'), is_admin, 
                     email or None, phone or None, address or None, logo_filename, notes or None, user_id)
                )
            else:
                conn.execute(
                    '''UPDATE users SET username=?, display_name=?, password_hash=?, is_admin=?, email=?, phone=?, address=?, notes=? 
                       WHERE id=?''',
                    (username, display_name or None, generate_password_hash(password, method='pbkdf2:sha256'), is_admin, 
                     email or None, phone or None, address or None, notes or None, user_id)
                )
        else:
            # 不更新密碼
            if logo_filename:
                conn.execute(
                    '''UPDATE users SET username=?, display_name=?, is_admin=?, email=?, phone=?, address=?, logo_file=?, notes=? 
                       WHERE id=?''',
                    (username, display_name or None, is_admin, email or None, phone or None, address or None, logo_filename, notes or None, user_id)
                )
            else:
                conn.execute(
                    '''UPDATE users SET username=?, display_name=?, is_admin=?, email=?, phone=?, address=?, notes=? 
                       WHERE id=?''',
                    (username, display_name or None, is_admin, email or None, phone or None, address or None, notes or None, user_id)
                )
        
        conn.commit()
    except Exception as e:
        logging.error(f"編輯使用者失敗: {e}")
    finally:
        conn.close()
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/users/<int:user_id>/toggle_active', methods=['POST'])
@admin_required
def admin_toggle_user_active(user_id: int):
    conn = get_db_connection()
    row = conn.execute('SELECT username, is_active FROM users WHERE id=?', (user_id,)).fetchone()
    if row:
        if row['username'] == 'admin':
            conn.close()
            return redirect(url_for('admin_dashboard'))
        new_active = 0 if row['is_active'] == 1 else 1
        conn.execute('UPDATE users SET is_active=? WHERE id=?', (new_active, user_id))
        conn.commit()
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def admin_delete_user(user_id: int):
    conn = get_db_connection()
    row = conn.execute('SELECT username FROM users WHERE id=?', (user_id,)).fetchone()
    if row and row['username'] != 'admin':
        conn.execute('DELETE FROM user_analyzers WHERE user_id=?', (user_id,))
        conn.execute('DELETE FROM users WHERE id=?', (user_id,))
        conn.commit()
    conn.close()
    return redirect(url_for('admin_dashboard'))

def _get_user_analyzer_settings_map(user_id: int) -> Dict[str, Dict[str, Any]]:
    conn = get_db_connection()
    rows = conn.execute('SELECT analyzer, enabled, vlm_provider, ollama_model, ocr_lang, save_files FROM user_analyzers WHERE user_id=?', (user_id,)).fetchall()
    conn.close()
    settings = {}
    for r in rows:
        settings[r['analyzer']] = {
            'enabled': bool(r['enabled']),
            'vlm_provider': r['vlm_provider'],
            'ollama_model': r['ollama_model'],
            'ocr_lang': r['ocr_lang'] or 'auto',  # 默認自動偵測
            'save_files': bool(r['save_files'])
        }
    return settings

@app.route('/admin/users/<int:user_id>/analyzers', methods=['GET', 'POST'])
@admin_required
def admin_user_analyzers(user_id: int):
    conn = get_db_connection()
    user = conn.execute('SELECT id, username FROM users WHERE id=?', (user_id,)).fetchone()
    conn.close()
    if not user:
        return redirect(url_for('admin_dashboard'))
    analyzers_list = list(config.get('prompts', {}).keys())
    if request.method == 'POST':
        conn = get_db_connection()
        for analyzer in analyzers_list:
            enabled = 1 if request.form.get(f'enabled_{analyzer}') == 'on' else 0
            provider = request.form.get(f'provider_{analyzer}', 'cloud')
            model = request.form.get(f'model_{analyzer}', None)
            ocr_lang = request.form.get(f'ocr_lang_{analyzer}', 'auto')
            save_files = 1 if request.form.get(f'save_files_{analyzer}') == 'on' else 0
            exist = conn.execute('SELECT id FROM user_analyzers WHERE user_id=? AND analyzer=?', (user_id, analyzer)).fetchone()
            if exist:
                conn.execute(
                    'UPDATE user_analyzers SET enabled=?, vlm_provider=?, ollama_model=?, ocr_lang=?, save_files=? WHERE user_id=? AND analyzer=?',
                    (enabled, provider, model, ocr_lang, save_files, user_id, analyzer)
                )
            else:
                conn.execute(
                    'INSERT INTO user_analyzers (user_id, analyzer, enabled, vlm_provider, ollama_model, ocr_lang, save_files) VALUES (?, ?, ?, ?, ?, ?, ?)',
                    (user_id, analyzer, enabled, provider, model, ocr_lang, save_files)
                )
        conn.commit()
        conn.close()
        return redirect(url_for('admin_user_analyzers', user_id=user_id) + '?saved=1')
    # GET
    settings = _get_user_analyzer_settings_map(user_id)
    return render_template('admin_user_analyzers.html', user=dict(user), analyzers=analyzers_list, settings=settings)

@app.route('/admin/history')
@admin_required
def admin_history():
    selected_username = request.args.get('username', '').strip()
    items = history
    if selected_username:
        items = [h for h in history if h.get('username') == selected_username]
    # 排序
    def get_history_time(x):
        return x.get('time') or x.get('timestamp') or ''
    items_sorted = sorted(items, key=get_history_time, reverse=True)
    # 需要 users 清單供下拉
    conn = get_db_connection()
    rows = conn.execute('SELECT id, username, display_name, is_admin, is_active, email, phone, address, logo_file, notes FROM users ORDER BY id ASC').fetchall()
    conn.close()
    users = [dict(r) for r in rows]
    
    # Create user mapping table (username -> user info)
    user_map = {u['username']: u for u in users}
    
    # Calculate statistics (for filtered records)
    stats = calculate_user_statistics(items_sorted)
    
    return render_template('admin_dashboard.html', users=users, history=items_sorted, user_map=user_map, selected_username=selected_username, stats=stats)

@app.route('/admin/history/delete', methods=['POST'])
@admin_required
def admin_delete_history():
    try:
        data = request.get_json()
        username = data.get('username')
        timestamp = data.get('timestamp')
        excel_file = data.get('excel_file')
        
        if not username or not timestamp:
            return jsonify({'success': False, 'error': 'Missing required parameters'})
        
        # Find and delete corresponding records from history
        global history
        original_count = len(history)
        
        # Find records to delete
        to_remove = []
        for i, h in enumerate(history):
            if (h.get('username') == username and 
                h.get('time') == timestamp):
                to_remove.append(i)
                
                # 刪除相關檔案
                if excel_file and h.get('excel') == excel_file:
                    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_file)
                    try:
                        if os.path.exists(excel_path):
                            os.remove(excel_path)
                            app.logger.info(f"刪除 Excel 檔案: {excel_path}")
                    except Exception as e:
                        app.logger.warning(f"刪除 Excel 檔案失敗: {e}")
                
                # 刪除日誌檔案
                log_file = h.get('log_file')
                if log_file:
                    log_path = os.path.join('logs', log_file)
                    try:
                        if os.path.exists(log_path):
                            os.remove(log_path)
                            app.logger.info(f"刪除日誌檔案: {log_path}")
                    except Exception as e:
                        app.logger.warning(f"刪除日誌檔案失敗: {e}")
        
        # Remove from history records (reverse order to avoid index issues)
        for i in sorted(to_remove, reverse=True):
            del history[i]
        
        # Save updated history records
        save_history()
        
        deleted_count = len(to_remove)
        if deleted_count > 0:
            app.logger.info(f"Administrator deleted {deleted_count} history records for user {username}")
            return jsonify({'success': True, 'message': f'Successfully deleted {deleted_count} records'})
        else:
            return jsonify({'success': False, 'error': 'No matching records found'})
            
    except Exception as e:
        app.logger.error(f"Failed to delete history records: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/history/clear-user', methods=['POST'])
@admin_required
def admin_clear_user_history():
    try:
        data = request.get_json()
        username = data.get('username')
        
        if not username:
            return jsonify({'success': False, 'error': 'Missing username parameter'})
        
        global history
        original_count = len(history)
        
        # Find records to delete
        to_remove = []
        for i, h in enumerate(history):
            if h.get('username') == username:
                to_remove.append(i)
                
                # Delete related files
                excel_file = h.get('excel')
                if excel_file:
                    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_file)
                    try:
                        if os.path.exists(excel_path):
                            os.remove(excel_path)
                            app.logger.info(f"Deleted Excel file: {excel_path}")
                    except Exception as e:
                        app.logger.warning(f"Failed to delete Excel file: {e}")
                
                # Delete log file
                log_file = h.get('log_file')
                if log_file:
                    log_path = os.path.join('logs', log_file)
                    try:
                        if os.path.exists(log_path):
                            os.remove(log_path)
                            app.logger.info(f"Deleted log file: {log_path}")
                    except Exception as e:
                        app.logger.warning(f"Failed to delete log file: {e}")
        
        # Remove from history records (reverse order to avoid index issues)
        for i in sorted(to_remove, reverse=True):
            del history[i]
        
        # Save updated history records
        save_history()
        
        deleted_count = len(to_remove)
        app.logger.info(f"Administrator cleared {deleted_count} history records for user {username}")
        return jsonify({'success': True, 'deleted_count': deleted_count})
            
    except Exception as e:
        app.logger.error(f"Failed to clear user history records: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/history/clear-all', methods=['POST'])
@admin_required
def admin_clear_all_history():
    try:
        global history
        original_count = len(history)
        
        # Delete all related files
        for h in history:
            # Delete Excel file
            excel_file = h.get('excel')
            if excel_file:
                excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_file)
                try:
                    if os.path.exists(excel_path):
                        os.remove(excel_path)
                        app.logger.info(f"Deleted Excel file: {excel_path}")
                except Exception as e:
                    app.logger.warning(f"Failed to delete Excel file: {e}")
            
            # Delete log file
            log_file = h.get('log_file')
            if log_file:
                log_path = os.path.join('logs', log_file)
                try:
                    if os.path.exists(log_path):
                        os.remove(log_path)
                        app.logger.info(f"Deleted log file: {log_path}")
                except Exception as e:
                    app.logger.warning(f"Failed to delete log file: {e}")
        
        # Clear all history records
        history.clear()
        save_history()
        
        app.logger.info(f"Administrator cleared all {original_count} history records")
        return jsonify({'success': True, 'deleted_count': original_count})
            
    except Exception as e:
        app.logger.error(f"Failed to clear all history records: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/ollama/models')
@admin_required
def admin_ollama_models():
    try:
        from analyzers.clients.ollama_client import get_ollama_models
        models = get_ollama_models()
        return jsonify({'success': True, 'models': models})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'models': []}), 500

@app.route('/admin/user-stats/<username>')
@admin_required
def admin_user_stats(username):
    """Get statistics for a specific user"""
    try:
        # Get the user's history records
        user_history = [h for h in history if h.get('username') == username]
        
        # Calculate statistics
        stats = calculate_user_statistics(user_history)
        
        if username in stats:
            return jsonify({
                'success': True,
                'stats': stats[username]
            })
        else:
            return jsonify({
                'success': True,
                'stats': {
                    'total_tokens': 0,
                    'total_records': 0,
                    'daily_usage': {}
                }
            })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
# 移除語言切換路由


if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port)


